import glob
import os
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)


pathToParentFolder = 'C:\\Users\\tarun\\Downloads\\'
folderName = "runs"
numTrials = 5

# data value to search for in config.txt
searchExpression = "strength=.+?(?=,)"
lengthOfSearch = 9

fullPath = pathToParentFolder + folderName

wb = Workbook()

ws = wb.active

runs = os.listdir(fullPath)

row = 2


strengths = {}

for run in runs:
    col = 1
    configFile = find("config.txt", fullPath + "\\" + run)
    detFile = find("det.txt", fullPath + "\\" + run)

    with open(configFile, "r") as file:
        text = file.read()
        regex = re.compile(searchExpression)
        result = regex.search(text)
        strength = float(
            text[result.span()[0]+lengthOfSearch: result.span()[1]])
        if strength not in strengths:
            strengths[strength] = {}

    with open(detFile, "r") as file:
        detections = strengths[strength]
        lines = file.readlines()
        for line in lines:
            data = line.split(":")
            detector = data[0]
            if detector not in detections:
                detections[detector] = []
            numDetections = data[1].strip()
            detections[detector].append(int(numDetections))

    #     ws[get_column_letter(col) + str(row)] = strength

    # for detector in detections:
    #     col += 1
    #     dataLocation = get_column_letter(col) + str(row)
    #     numDetections = detections[detector]
    #     ws[dataLocation] = float(sum(numDetections) / len(numDetections))

    # row += 1

# print(strengths)

ws['A1'] = "Strength"
col = 1
for detector in strengths[0.0]:
    col += 1
    ws[get_column_letter(col) + '1'] = detector


row = 2
for strength in strengths:
    col = 1
    ws[get_column_letter(col) + str(row)] = strength
    allDetectors = strengths[strength]
    for detector in allDetectors:
        col += 1
        dataLocation = get_column_letter(col) + str(row)
        detections = allDetectors[detector]
        ws[dataLocation] = float(sum(detections) / len(detections))

    row += 1

wb.save(filename="Results.xlsx")
wb.close()
