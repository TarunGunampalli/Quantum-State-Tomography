import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font


def checkValidDensityMatrix(dm, input, probType):
    hermitian = np.array_equal(dm, dm.getH())
    trace1 = (dm.item(0, 0) + dm.item(1, 1)) == 1
    eigen1 = 0.5 + np.sqrt(0.25 - ((dm.item(0, 0) *
                                    dm.item(1, 1)) - (dm.item(0, 1) * dm.item(1, 0))))
    eigen2 = 0.5 - np.sqrt(0.25 - ((dm.item(0, 0) *
                                    dm.item(1, 1)) - (dm.item(0, 1) * dm.item(1, 0))))
    validEigens = eigen1.real >= 0 and eigen2.real >= 0 and eigen1.imag == 0 and eigen2.imag == 0

    # if not hermitian:
    #     print(str(dm) + "(" + probType +
    #           ") is not hermitian for data input " + str(input))
    # if not trace1:
    #     print(str(dm) + "(" + probType +
    #           ") is does not have a trace of 1 for data input " + str(input))
    # if not validEigens:
    #     print(str(dm) + "(" + probType +
    #           ") does not have valid eigenvalues for data input " + str(input))
    #     print(eigen1.real)
    #     print(eigen2.real)

    return hermitian and trace1 and validEigens


def calculateFidelities(ws):
    inputs = []

    EZ = {}
    EZRaw = {}
    EZAdj = {}

    for row in ws.iter_rows(min_row=3,  max_row=171, min_col=1, max_col=9, values_only=True):
        inputParams = (row[0], row[1])
        inputs.append(inputParams)
        expected = row[6]
        expectedRaw = row[7]
        expectedAdj = row[8]

        EZ[inputParams] = expected
        EZRaw[inputParams] = expectedRaw
        EZAdj[inputParams] = expectedAdj

    EX = {}
    EXRaw = {}
    EXAdj = {}

    for row in ws.iter_rows(min_row=3,  max_row=171, min_col=11, max_col=19, values_only=True):
        inputParams = (row[0], row[1])
        expected = row[6]
        expectedRaw = row[7]
        expectedAdj = row[8]

        EX[inputParams] = expected
        EXRaw[inputParams] = expectedRaw
        EXAdj[inputParams] = expectedAdj

    EY = {}
    EYRaw = {}
    EYAdj = {}

    for row in ws.iter_rows(min_row=3,  max_row=171, min_col=21, max_col=29, values_only=True):
        inputParams = (row[0], row[1])
        expected = row[6]
        expectedRaw = row[7]
        expectedAdj = row[8]

        EY[inputParams] = expected
        EYRaw[inputParams] = expectedRaw
        EYAdj[inputParams] = expectedAdj

    I = np.matrix([[1, 0],
                   [0, 1]])

    Z = np.matrix([[1, 0],
                   [0, -1]])

    X = np.matrix([[0, 1],
                   [1, 0]])

    Y = np.matrix([[0, complex(0, -1)],
                   [complex(0, 1), 0]])

    initialState = np.matrix([[1],
                              [0]])

    # newWS = wb.create_sheet(title="Fidelities")
    newWS = ws
    newWS['AE2'] = "Fidelity Adjusted"
    newWS['AF2'] = "Fidelity Raw"
    newWS['AG2'] = "Fidelity Adjusted 2"
    row = 3

    bold = Font(bold=True)

    for input in inputs:
        actual = 0.5 * (1 * I + EZ[input] * Z + EX[input] * X + EY[input] * Y)
        actualRaw = 0.5 * (1 * I + EZRaw[input] *
                           Z + EXRaw[input] * X + EYRaw[input] * Y)
        actualAdj = 0.5 * (1 * I + EZAdj[input] *
                           Z + EXAdj[input] * X + EYAdj[input] * Y)

        valid = checkValidDensityMatrix(actual, input, "adjusted")
        validRaw = checkValidDensityMatrix(actualRaw, input, "raw")
        validAdj = checkValidDensityMatrix(
            actualAdj, input, "adjusted with single dark count")

        thetaHWP = np.radians(input[0])
        HWP = np.matrix([[np.cos(2 * thetaHWP), np.sin(2 * thetaHWP)],
                         [np.sin(2 * thetaHWP), -np.cos(2 * thetaHWP)]])

        thetaQWP = np.radians(input[1])
        QWP = np.matrix([[np.square(np.cos(thetaQWP)) + complex(0, 1) * np.square(np.sin(thetaQWP)), complex(1, -1) * (np.cos(thetaQWP) * np.sin(thetaQWP))],
                         [complex(1, -1) * (np.cos(thetaQWP) * np.sin(thetaQWP)), np.square(np.sin(thetaQWP)) + complex(0, 1) * np.square(np.cos(thetaQWP))]])

        psi = QWP * HWP * initialState

        fidelity = psi.getH() * actual * psi
        fidelityRaw = psi.getH() * actualRaw * psi
        fidelityAdj = psi.getH() * actualAdj * psi
        newWS['AE' + str(row)] = fidelity.real.item(0, 0)
        if not valid:
            newWS['AE' + str(row)].font = bold
        newWS['AF' + str(row)] = fidelityRaw.real.item(0, 0)
        if not validRaw:
            newWS['AF' + str(row)].font = bold
        newWS['AG' + str(row)] = fidelityAdj.real.item(0, 0)
        if not validAdj:
            newWS['AG' + str(row)].font = bold
        row += 1


wb = load_workbook("Quantum Computing Research.xlsx", data_only=True)

sheets = ["Quantum State Tomography", "OD 9.9", "OD 10.1", "DC 90", "DC 110"]

for sheet in sheets:
    ws = wb[sheet]
    calculateFidelities(ws)

wb.save(filename="Quantum Computing Research.xlsx")
wb.close()
