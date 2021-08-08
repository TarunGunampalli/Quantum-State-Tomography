import glob
import os
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


###############################################
###############################################
###############################################
############# VARIABLES TO CHANGE #############
###############################################
###############################################
###############################################

pathToParentFolder = 'C:\\Users\\tarun\\Downloads\\'
folderName = "runs"

# data values to search for in config.txt
# key is the name of the variable
# value is a tuple
# index 0 being the regex to search for in det.txt
# index 1 is the length between the start of the regex and the position of the value
variables = {
    'angle': ("angle=.+?(?=\n)", len('angle=')),
    # 'phase': ("phase=.+?(?=\n)", len('phase='))
}

resultsFileName = '../Data/Results.xlsx'

###############################################
###############################################
###############################################
###############################################
###############################################
###############################################
###############################################


fullPath = pathToParentFolder + folderName

wb = Workbook()

ws = wb.active

runs = os.listdir(fullPath)

inputs = {}


def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)


variableNames = []
namesRead = False
# Read inputs from config file and data from det.txt and add inputs and outputs to a dictionary
for run in runs:
    col = 1
    configFile = find("config.txt", fullPath + "\\" + run)
    detFile = find("det.txt", fullPath + "\\" + run)

    with open(configFile, "r") as file:
        text = file.read()
        input = []
        for variable in variables:
            if not namesRead:
                variableNames.append(variable)
            searchExpression = variables[variable][0]
            length = variables[variable][1]
            regex = re.compile(searchExpression)
            result = regex.search(text)
            numInstances = 1
            while result:
                textStart = result.span()[0]+length
                textEnd = result.span()[1]
                input.append(float(text[textStart: textEnd]))
                # print(result)
                text = text[textEnd:]
                result = regex.search(text)
                if numInstances > 1 and not namesRead:
                    variableNames.append(variable + str(numInstances))
                numInstances += 1

        input = tuple(input)
        if input not in inputs:
            inputs[input] = {}

    namesRead = True

    with open(detFile, "r") as file:
        detections = inputs[input]
        lines = file.readlines()
        for line in lines:
            data = line.split(":")
            detector = data[0]
            if detector not in detections:
                detections[detector] = []
            numDetections = data[1].strip()
            detections[detector].append(int(numDetections))


row = 1
col = 1
# add input names to excel sheet
for variable in variableNames:
    ws[get_column_letter(col) + str(row)] = variable
    col += 1

col = len(variableNames)
# add detector combinations to excel sheet
for detector in inputs[list(inputs.keys())[0]]:
    col += 1
    ws[get_column_letter(col) + str(row)] = detector


row = 2
# add the input values and data values to the excel sheet
for input in inputs:
    col = 0
    for variable in input:
        col += 1
        ws[get_column_letter(col) + str(row)] = variable
    allDetectors = inputs[input]
    for detector in allDetectors:
        col += 1
        dataLocation = get_column_letter(col) + str(row)
        detections = allDetectors[detector]
        ws[dataLocation] = float(sum(detections) / len(detections))

    row += 1

wb.save(filename=resultsFileName)
wb.close()
