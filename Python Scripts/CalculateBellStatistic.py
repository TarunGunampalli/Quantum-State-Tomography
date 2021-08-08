from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

pathToParentFolder = 'C:\\Users\\tarun\\OneDrive - The University of Texas at Austin\\Documents\\School\\CS 309\\Data\\'
excelName = "Bell's Inequality.xlsx"

fullPath = pathToParentFolder + excelName

wb = load_workbook(fullPath, data_only=True)

ws = wb['Expected Values']

exColumn = column_index_from_string('T')

expectedValues = {}

for row in ws.iter_rows(min_row=2, max_row=101, min_col=0, max_col=exColumn):
    angle1 = row[0].value
    angle2 = row[1].value
    expectedValue = row[exColumn - 1].value

    expectedValues[(angle1, angle2)] = expectedValue

# print(expectedValues)

bellStatistics = []
angles = {}

for A in range(0, 100, 10):
    for B in range(0, 100, 10):
        for AP in range(0, 100, 10):
            for BP in range(0, 100, 10):
                ex1 = expectedValues[(A, B)]
                ex2 = expectedValues[(A, BP)]
                ex3 = expectedValues[(AP, B)]
                ex4 = expectedValues[(AP, BP)]

                bellStatistic = abs(ex1 + ex2) + abs(ex3 - ex4)
                bellStatistics.append(bellStatistic)
                angles[bellStatistic] = (A, B, AP, BP)

maxBS = max(bellStatistics)

print(f'Max Bell Statistic = {maxBS}')
print(f'Input Angles = {angles[maxBS]}')
