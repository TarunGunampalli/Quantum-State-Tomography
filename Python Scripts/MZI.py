import glob
import os
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)


pathToParentFolder = 'C:\\Users\\tarun\\Downloads\\'
folderName = "runs"

# data value to search for in config.txt
searchExpression = "phase=.+?(?=\n)"
lengthOfSearch = len("phase=")

fullPath = pathToParentFolder + folderName

wb = Workbook()

ws = wb.active

runs = os.listdir(fullPath)

row = 2


phases = {}
# phases2 = {}

for run in runs:
    col = 1
    configFile = find("config.txt", fullPath + "\\" + run)
    detFile = find("pm.csv", fullPath + "\\" + run)

    with open(configFile, "r") as file:
        text = file.read()
        regex = re.compile(searchExpression)
        result = regex.search(text)
        phase = float(
            text[result.span()[0]+lengthOfSearch: result.span()[1]])
        phase = round(phase, 2)

    with open(detFile, "r") as file:
        text = file.read()
        regex = re.compile("1 ,.+?(?=,)")
        result = regex.search(text)
        power = float(
            text[result.span()[0]+3: result.span()[1]])
        phases[phase] = power

        # regex = re.compile("\d+\.\d+,.+?(?=\n)")
        # result = regex.search(text)
        # print(text[result.span()[0] + 3: result.span()[1]].split(",")[1])
        # power = float(
        #     text[result.span()[0] + 3: result.span()[1]].split(","))

        # phases2[phase] = power

ws['A1'] = "Phases"
col = 2

# print(phases)


row = 2
for phase in phases:
    col = 1
    ws[get_column_letter(col) + str(row)] = phase
    col = 2
    dataLocation = get_column_letter(col) + str(row)
    ws[dataLocation] = phases[phase]

    row += 1

wb.save(filename="Results.xlsx")
wb.close()
