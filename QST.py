from openpyxl import load_workbook
import numpy as np

wb = load_workbook("Book1.xlsx", data_only=True)

ws = wb.active

inputs = []

EZ = {}
EZRaw = {}
EZadj2 ={}

for row in ws.iter_rows(min_row=3,  max_row=171, min_col=1, max_col=9, values_only=True):
    inputParams = (row[0], row[1])
    inputs.append(inputParams)
    expected = row[6]
    expectedRaw = row[7]
    expected_adj2 = row[8]

    EZ[inputParams] = expected
    EZRaw[inputParams] = expectedRaw
    EZadj2[inputParams] = expected_adj2

EX = {}
EXRaw = {}
EXadj2 ={}

for row in ws.iter_rows(min_row=3,  max_row=171, min_col=11, max_col=19, values_only=True):
    inputParams = (row[0], row[1])
    expected = row[6]
    expectedRaw = row[7]
    expected_adj2 = row[8]

    EX[inputParams] = expected
    EXRaw[inputParams] = expectedRaw
    EXadj2[inputParams] = expected_adj2

EY = {}
EYRaw = {}
EYadj2 ={}

for row in ws.iter_rows(min_row=3,  max_row=171, min_col=21, max_col=29, values_only=True):
    inputParams = (row[0], row[1])
    expected = row[6]
    expectedRaw = row[7]
    expected_adj2 = row[8]

    EY[inputParams] = expected
    EYRaw[inputParams] = expectedRaw
    EYadj2[inputParams] = expected_adj2

I = np.matrix([[1, 0],
               [0, 1]])

Z = np.matrix([[1, 0],
               [0, -1]])

X = np.matrix([[0, 1],
               [1, 0]])

Y = np.matrix([[0, complex(0, -1)],
               [complex(0, 1), 0]])

initialState = np.matrix([[1],
                          [0]])

newWS = wb.create_sheet(title="Fidelities")
newWS['A1'] = "Fidelity adj"
newWS['B1'] = "Fidelity Raw"
newWS['C1'] = "Fidelity adj2"
row = 3

for input in inputs:
    actual = 0.5 * (1 * I + EZ[input] * Z + EX[input] * X + EY[input] * Y)
    actualRaw = 0.5 * (1 * I + EZRaw[input] *
                       Z + EXRaw[input] * X + EYRaw[input] * Y)
    actualAdj = 0.5 * (1 * I + EZadj2[input] *
                       Z + EXadj2[input] * X + EYadj2[input] * Y)

    thetaHWP = np.radians(input[0])
    HWP = np.matrix([[np.cos(2 * thetaHWP), np.sin(2 * thetaHWP)],
                     [np.sin(2 * thetaHWP), -np.cos(2 * thetaHWP)]])

    thetaQWP = np.radians(input[1])
    QWP = np.matrix([[np.square(np.cos(thetaQWP)) + complex(0, 1) * np.square(np.sin(thetaQWP)), complex(1, -1) * (np.cos(thetaQWP) * np.sin(thetaQWP))],
                     [complex(1, -1) * (np.cos(thetaQWP) * np.sin(thetaQWP)), np.square(np.sin(thetaQWP)) + complex(0, 1) * np.square(np.cos(thetaQWP))]])

    psi = QWP * HWP * initialState

    fidelity = psi.getH() * actual * psi
    fidelityRaw = psi.getH() * actualRaw * psi
    fidelityadj = psi.getH() * actualAdj * psi
    newWS['A' + str(row)] = fidelity.real.item(0, 0)
    newWS['B' + str(row)] = fidelityRaw.real.item(0, 0)
    newWS['C' + str(row)] = fidelityadj.real.item(0, 0)
    row += 1

wb.save(filename="Book1.xlsx")
