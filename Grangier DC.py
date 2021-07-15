import glob
import os
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

######### VARIABLES TO CHANGE #########
pathToParentFolder = 'C:\\Users\\tarun\\Downloads'
folderName = "runs"

# data value to search for in config.txt
searchExpression = "optical_density=.+?(?=\n)"
# "strength=.+?(?=,)"
# "dark_count_rate=\d+?(?=\n)"
lengthOfSearch = len("optical_density=")
# len("strength=")
# len("dark_count_rate=")


def castValue(val):
    return float(val)  # replace with type of the data value


# title of independent variable
title = "Optical Density"
######### END VARIABLES TO CHANGE #########


def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)


fullPath = f"{pathToParentFolder}\\{folderName}"

wb = Workbook()

ws = wb.active

runs = os.listdir(fullPath)

row = 2


dcRates = {}

for run in runs:
    col = 1
    configFile = find("config.txt", f"{fullPath}\\{run}")
    detFile = find("det.txt", f"{fullPath}\\{run}")

    with open(configFile, "r") as file:
        text = file.read()
        regex = re.compile(searchExpression)
        result = regex.search(text)
        dcRate = text[result.span()[0]+lengthOfSearch: result.span()[1]]
        dcRate = castValue(dcRate)
        if dcRate not in dcRates:
            dcRates[dcRate] = {}

    with open(detFile, "r") as file:
        detections = dcRates[dcRate]
        lines = file.readlines()
        for line in lines:
            data = line.split(":")
            detector = data[0]
            if detector not in detections:
                detections[detector] = []
            numDetections = data[1].strip()
            detections[detector].append(int(numDetections))

    #     ws[get_column_letter(col) + str(row)] = strength

    # for detector in detections:
    #     col += 1
    #     dataLocation = get_column_letter(col) + str(row)
    #     numDetections = detections[detector]
    #     ws[dataLocation] = float(sum(numDetections) / len(numDetections))

    # row += 1

# print(dcRates.keys())

ws['A1'] = title
col = 1
for detector in dcRates[list(dcRates.keys())[0]]:
    col += 1
    ws[get_column_letter(col) + '1'] = detector


row = 2
for dcRate in dcRates:
    col = 1
    ws[get_column_letter(col) + str(row)] = dcRate
    allDetectors = dcRates[dcRate]
    for detector in allDetectors:
        col += 1
        dataLocation = get_column_letter(col) + str(row)
        detections = allDetectors[detector]
        ws[dataLocation] = float(sum(detections) / len(detections))

    row += 1

wb.save(filename="Results.xlsx")
wb.close()
